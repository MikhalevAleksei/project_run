from django.db.models import Sum
from geopy.distance import geodesic

from django.contrib.auth.models import User
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import load_workbook
from rest_framework.filters import OrderingFilter
from rest_framework.pagination import PageNumberPagination
from django.shortcuts import get_object_or_404
from rest_framework.decorators import api_view
from rest_framework.filters import SearchFilter
from rest_framework.response import Response
from django.conf import settings
from rest_framework import viewsets, status, generics
from rest_framework.views import APIView

from .models import Run, AthleteInfo, Challenge, Position, CollectibleItem
from .serializers import (
    RunSerializer,
    UserSerializer,
    AthleteInfoSerializer,
    ChallengeSerializer,
    PositionSerializer, CollectibleItemSerializer,
)
from .helpers import calculate_run_distance


class StandardPagination(PageNumberPagination):
    # page_size = 10  # Количество объектов на странице по умолчанию (не обязательный параметр)
    page_size_query_param = 'size'
    # max_page_size = 100  # Ограничиваем максимальное количество объектов на странице


class RunViewSet(viewsets.ModelViewSet):
    queryset = Run.objects.select_related('athlete').all()

    serializer_class = RunSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['status', 'athlete']
    ordering_fields = ['created_at']
    ordering = ['id']
    pagination_class = StandardPagination


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = UserSerializer
    queryset = User.objects.all()
    filter_backends = [SearchFilter, OrderingFilter]  # Подключаем SearchFilter здесь
    search_fields = ['first_name', 'last_name']  # Указываем поля по которым будет вестись поиск
    ordering_fields = ['date_joined']
    ordering = ['id']
    pagination_class = StandardPagination

    def get_queryset(self):
        qs = self.queryset.exclude(is_superuser=True)  # исключаем админов
        user_type = self.request.query_params.get("type")

        if user_type == "coach":
            qs = qs.filter(is_staff=True)
        elif user_type == "athlete":
            qs = qs.filter(is_staff=False)

        return qs


@api_view(['GET'])
def company_details(request):
    details = {'company_name': settings.COMPANY_NAME,
               'slogan': settings.SLOGAN,
               'contacts': settings.CONTACTS}
    return Response(details)


class StartStatusAPIView(APIView):
    def get(self, request, run_id, *args, **kwargs):
        run = get_object_or_404(Run, id=run_id)
        return Response(
            {"status": run.status},
            status=status.HTTP_200_OK,
        )

    def post(self, request, run_id, *args, **kwargs):
        run = get_object_or_404(Run, id=run_id)

        if run.status != 'init':
            return Response(
                {"error": "Забег уже начат или завершён."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        run.status = 'in_progress'
        run.save()
        return Response(
            {"message": "Забег запущен", "status": run.status},
            status=status.HTTP_200_OK,
        )


class StopStatusAPIView(APIView):
    def get(self, request, run_id, *args, **kwargs):
        run = get_object_or_404(Run, id=run_id)
        return Response(
            {"status": run.status},
            status=status.HTTP_200_OK,
        )

    def post(self, request, run_id, *args, **kwargs):
        run = get_object_or_404(Run, id=run_id)

        if run.status != Run.STATUS_IN_PROGRESS:
            return Response(
                {"error": "Нельзя завершить забег, который не запущен или уже завершён."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        run.distance = calculate_run_distance(run)
        run.status = Run.STATUS_FINISHED
        run.save()

        challenge_count = Run.objects.filter(
            athlete=run.athlete,
            status=Run.STATUS_FINISHED,
        ).count()

        if challenge_count == 10:
            Challenge.objects.create(athlete=run.athlete)

        total_distance = Run.objects.filter(
            athlete=run.athlete,
            status=Run.STATUS_FINISHED,
        ).aggregate(
            total_distance=Sum('distance')
        )

        athlete_total_distance = total_distance.get('total_distance') or 0

        if athlete_total_distance >= 50:
            Challenge.objects.get_or_create(
                athlete=run.athlete,
                title="Пробеги 50 километров!",
            )

        return Response(
            {
                "message": "Забег завершён",
                "status": run.status,
                "distance": run.distance,
            },
            status=status.HTTP_200_OK,
        )


class AthleteInfoAPIView(APIView):
    """
    GET /api/athlete_info/<user_id>/
    PUT /api/athlete_info/<user_id>/
    """

    def get(self, request, user_id):
        # Проверяем, существует ли User
        user = get_object_or_404(User, id=user_id)

        # Получаем или создаём AthleteInfo
        athlete_info, created = AthleteInfo.objects.get_or_create(user=user)

        serializer = AthleteInfoSerializer(athlete_info)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self, request, user_id):
        # Проверяем, существует ли User
        user = get_object_or_404(User, id=user_id)

        # Получаем данные
        goals = request.data.get('goals')
        weight = request.data.get('weight')

        # Проверяем корректность веса
        if weight is not None:
            try:
                weight = int(weight)
            except ValueError:
                return Response(
                    {"error": "weight должен быть числом"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if weight <= 0 or weight >= 900:
                return Response(
                    {"error": "weight должен быть больше 0 и меньше 900"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # Создаём или обновляем запись
        athlete_info, created = AthleteInfo.objects.update_or_create(
            user=user,
            defaults={
                "goals": goals,
                "weight": weight,
            },
        )

        serializer = AthleteInfoSerializer(athlete_info)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


# class ChallengesAPIView(APIView):
#     """
#     GET /api/challenges/
#     GET /api/challenges/?athlete=<id>
#     """
#     def get(self, request):
#         athlete_id = request.query_params.get("athlete")
#         queryset = Challenges.objects.all()
#
#         if athlete_id:
#             queryset = queryset.filter(athlete_id=athlete_id)
#
#         serializer = ChallengesSerializer(queryset, many=True)
#         return Response(serializer.data, status=status.HTTP_200_OK)

class ChallengeViewSet(viewsets.ReadOnlyModelViewSet):
    """
    /api/challenges/ - список всех челленджей
    /api/challenges/?athlete=<id> - фильтр по пользователю
    /api/challenges/?ordering=-created_at - сортировка по дате
    /api/challenges/?page=2&size=5 - пагинация
    """
    queryset = Challenge.objects.all()
    serializer_class = ChallengeSerializer
    pagination_class = StandardPagination

    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['athlete']
    ordering_fields = ['created_at', ]
    ordering = ['-created_at']  # сортировка по умолчанию (новые сверху)


class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer

    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['run']

    http_method_names = ['get', 'post', 'delete', 'head', 'options']


class CollectibleItemListView(generics.ListAPIView):
    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer


class UploadFileView(APIView):

    def post(self, request):
        uploaded_file = request.FILES.get("file")

        if not uploaded_file:
            return Response(
                {"error": "File is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        workbook = load_workbook(uploaded_file)
        sheet = workbook.active

        invalid_rows = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = {
                "name": row[0],
                "uid": row[1],
                "value": row[2],
                "latitude": row[3],
                "longitude": row[4],
                "picture": row[5],
            }

            serializer = CollectibleItemSerializer(data=data)

            if serializer.is_valid():
                serializer.save()
            else:
                invalid_rows.append(list(row))

        return Response(
            invalid_rows,
            status=status.HTTP_200_OK,
        )

